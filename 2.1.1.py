namesForstatForYear = ["Год","Средняя зарплата",	"Средняя зарплата - Программист",	"Количество вакансий",	"Количество вакансий - Программист"]
namesForStatForCity =["Город","Уровень зарплат", "Доля вакансий"]
dictSalaryForYear = {2007: 38916, 2008: 43646, 2009: 42492, 2010: 43846, 2011: 47451, 2012: 48243, 2013: 51510, 2014: 50658}
dictSalaryForYearForProf = {2007: 43770, 2008: 50412, 2009: 46699, 2010: 50570, 2011: 55770, 2012: 57960, 2013: 58804, 2014: 62384}
dictCountVacansy= {2007: 2196, 2008: 17549, 2009: 17709, 2010: 29093, 2011: 36700, 2012: 44153, 2013: 59954, 2014: 66837}
dictCountVacansyForProf ={2007: 317, 2008: 2460, 2009: 2066, 2010: 3614, 2011: 4422, 2012: 4966, 2013: 5990, 2014: 5492}
dictLevelSalaryForYearForCity = {"Москва": 57354, "Санкт-Петербург": 46291, "Новосибирск": 41580, "Екатеринбург": 41091, "Казань": 37587, "Самара": 34091, "Нижний Новгород": 33637, "Ярославль": 32744, "Краснодар": 32542, "Воронеж": 29725}
dictPercentVacancyForcity = {"Москва": 0.4581, "Санкт-Петербург": 0.1415, "Нижний Новгород": 0.0269, "Казань": 0.0266, "Ростов-на-Дону": 0.0234, "Новосибирск": 0.0202, "Екатеринбург": 0.0143, "Воронеж": 0.014, "Самара": 0.0133, "Краснодар": 0.0131}


from openpyxl import Workbook
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Font
wb = Workbook()

statForYear = wb.active
statForYear.title = "Статистика по годам"
statForCity = wb.create_sheet("Статистика по городам")

statForCity['A1'] = 42
for i, statN in enumerate(namesForstatForYear):
    statForYear.cell(row=1, column=i+1).value = statN
    statForYear.cell(row=1, column=i + 1).font = Font(bold=True)
for i, statN in enumerate(dictSalaryForYear.keys()):
    statForYear.cell(row=i+2, column=1).value = statN
for i, statN in enumerate(dictSalaryForYear.values()):
    statForYear.cell(row=i+2, column=2).value = statN
for i, statN in enumerate(dictSalaryForYearForProf.values()):
    statForYear.cell(row=i + 2, column=3).value = statN
for i, statN in enumerate(dictCountVacansy.values()):
    statForYear.cell(row=i + 2, column=4).value = statN
for i, statN in enumerate(dictCountVacansyForProf.values()):
    statForYear.cell(row=i + 2, column=5).value = statN

for i, statN in enumerate(namesForStatForCity[0:2]):
        statForCity.cell(row=1, column=i+1).value = statN
        statForCity.cell(row=1, column=i + 1).font = Font(bold=True)
for i, statN in enumerate(namesForStatForCity[0::2]):
        statForCity.cell(row=1, column=i+4).value = statN
        statForCity.cell(row=1, column=i + 4).font = Font(bold=True)
for i, statN in enumerate(dictLevelSalaryForYearForCity.keys()):
        statForCity.cell(row=i+2, column=1).value = statN
for i, statN in enumerate(dictLevelSalaryForYearForCity.values()):
        statForCity.cell(row=i+2, column=2).value = statN
for i, statN in enumerate(dictPercentVacancyForcity.keys()):
        statForCity.cell(row=i+2, column=4).value = statN
for i, statN in enumerate(dictPercentVacancyForcity.values()):
        statForCity.cell(row=i + 2, column=5).value = statN
        statForCity.cell(row=i + 2, column=5).number_format = '0.00%'


thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

for i in range(9):
    for j in range(5):
        statForYear.cell(row=i+1, column=j+1).border = thin_border

for i in range(11):
    for j in range(5):
        if j!=2:
            statForCity.cell(row=i+1, column=j+1).border = thin_border


def padding(list):
    column_widths = []
    for row in list.iter_rows():
        for i, cell in enumerate(row):
            try:
                column_widths[i] = max(column_widths[i], len(str(cell.value)))
            except IndexError:
                column_widths.append(len(str(cell.value)))

    for i, column_width in enumerate(column_widths):
        list.column_dimensions[get_column_letter(i + 1)].width = column_width+1
padding(statForYear)
padding(statForCity)

wb.save("sample.xlsx")
